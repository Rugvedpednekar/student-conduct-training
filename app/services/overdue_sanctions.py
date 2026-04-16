from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO
from typing import Dict, Iterable, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill(fill_type="solid", fgColor="2D4A6B")
HEADER_FONT = Font(color="FFFFFF", bold=True)
MISSING_EMAIL_FILL = PatternFill(fill_type="solid", fgColor="FFE7E7")
MISSING_DEADLINE_FILL = PatternFill(fill_type="solid", fgColor="FFF5D6")

REQUIRED_OUTPUT_FIELDS = [
    "Full Name",
    "Student ID",
    "Email",
    "Assigned Sanctions",
    "Next Deadline",
    "Next Deadline Reason",
    "Hold",
    "Send an Email",
]

PREVIEW_ROW_LIMIT = 25

HEADER_ALIASES = {
    "file_id": {"fileid", "file_id", "file id", "case id"},
    "full_name": {"full name", "student name", "name", "respondent name"},
    "first_name": {"first name", "firstname", "given name"},
    "last_name": {"last name", "lastname", "surname", "family name"},
    "student_id": {"student id", "sid", "id", "studentid"},
    "email": {"email", "email address", "student email"},
    "assigned_sanctions": {
        "assigned sanctions",
        "sanctions",
        "sanctions/actions",
        "sanctions actions",
        "actions",
        "assigned actions",
        "assigned sanction",
    },
    "next_deadline": {"next deadline", "deadline", "upcoming deadline"},
    "next_deadline_reason": {"next deadline reason", "deadline reason", "deadline notes"},
    "status": {"status", "case status", "student status"},
}


class OverdueSanctionsError(Exception):
    pass


@dataclass
class ProcessedOverdueSanctionsResult:
    total_rows: int
    preview_columns: List[str]
    preview_rows: List[Dict[str, str]]
    workbook_content: bytes


def _normalize_header(value: object) -> str:
    text = "" if value is None else str(value)
    lowered = " ".join(text.strip().lower().replace("_", " ").split())
    return "".join(ch for ch in lowered if ch.isalnum() or ch == " ").strip()


def _is_missing(value: object) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def _safe_cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def _parse_date_value(value: object) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        raw = value.strip()
        if not raw:
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%b %d, %Y", "%B %d, %Y"):
            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                continue
    return None


def _normalize_phrase(value: str) -> str:
    return " ".join("".join(ch for ch in value.lower().strip() if ch.isalnum() or ch.isspace()).split())


def _is_student_account_hold(status: str) -> bool:
    normalized = _normalize_phrase(status)
    return normalized in {"student account hold", "student acct hold"} or (
        "student" in normalized and "hold" in normalized and ("account" in normalized or "acct" in normalized)
    )


def _find_column_indexes(ws: Worksheet) -> Dict[str, int]:
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not first_row:
        raise OverdueSanctionsError("The uploaded workbook is empty.")

    normalized_map = {}
    for idx, header in enumerate(first_row, start=1):
        key = _normalize_header(header)
        if key:
            normalized_map[key] = idx

    resolved: Dict[str, int] = {}
    for canonical, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            alias_key = _normalize_header(alias)
            if alias_key in normalized_map:
                resolved[canonical] = normalized_map[alias_key]
                break

    if "assigned_sanctions" not in resolved:
        raise OverdueSanctionsError(
            "Missing sanctions/actions column. Please export the overdue sanctions report from Maxient and try again."
        )

    if "full_name" not in resolved and ("first_name" not in resolved or "last_name" not in resolved):
        raise OverdueSanctionsError(
            "Missing student name columns. Include Full Name or both First Name and Last Name in the export."
        )
    if "status" not in resolved:
        raise OverdueSanctionsError("Missing required column: Status.")

    return resolved


def _build_output_row(row: tuple, columns: Dict[str, int]) -> Dict[str, str]:
    def pick(column_name: str) -> str:
        idx = columns.get(column_name)
        if not idx:
            return ""
        return _safe_cell_text(row[idx - 1])

    full_name = pick("full_name")
    if not full_name:
        full_name = " ".join(part for part in [pick("first_name"), pick("last_name")] if part).strip()

    status_value = pick("status")
    hold_value = "Yes" if _is_student_account_hold(status_value) else "No"
    send_email_value = "Yes" if hold_value == "Yes" else "No"

    return {
        "Full Name": full_name,
        "Student ID": pick("student_id"),
        "Email": pick("email"),
        "Assigned Sanctions": pick("assigned_sanctions"),
        "Next Deadline": pick("next_deadline"),
        "Next Deadline Reason": pick("next_deadline_reason"),
        "Hold": hold_value,
        "Send an Email": send_email_value,
    }


def _set_column_widths(ws: Worksheet, max_width: int = 52) -> None:
    for col_idx, column_cells in enumerate(ws.iter_cols(min_row=1, max_col=ws.max_column, max_row=ws.max_row), start=1):
        measured_width = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            longest_line = max((len(line) for line in value.splitlines()), default=0)
            measured_width = max(measured_width, longest_line)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, measured_width + 2), max_width)


def _append_summary_sheet(workbook: Workbook, records: Iterable[Dict[str, str]]) -> None:
    records = list(records)
    summary = workbook.create_sheet(title="Summary")

    total_rows = len(records)
    missing_email = sum(1 for item in records if _is_missing(item.get("Email")))
    missing_deadline = sum(1 for item in records if _is_missing(item.get("Next Deadline")))
    hold_yes = sum(1 for item in records if item.get("Hold", "").strip().lower() == "yes")
    send_email_yes = sum(1 for item in records if item.get("Send an Email", "").strip().lower() == "yes")

    rows = [
        ("Metric", "Count"),
        ("Total rows", total_rows),
        ("Missing email", missing_email),
        ("Missing next deadline", missing_deadline),
        ("Hold = Yes", hold_yes),
        ("Send an Email = Yes", send_email_yes),
    ]

    for row in rows:
        summary.append(row)

    for cell in summary[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

    summary.column_dimensions["A"].width = 36
    summary.column_dimensions["B"].width = 14


def process_overdue_sanctions_workbook(content: bytes, *, date_from: date, date_to: date) -> ProcessedOverdueSanctionsResult:
    if date_from > date_to:
        raise OverdueSanctionsError("Invalid date range. 'From' date cannot be later than 'To' date.")

    try:
        source_wb = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as exc:
        raise OverdueSanctionsError("Invalid workbook. Please upload a valid .xlsx file exported from Maxient.") from exc

    source_ws = source_wb.active
    columns = _find_column_indexes(source_ws)
    if "next_deadline" not in columns:
        raise OverdueSanctionsError("Missing required column: Next Deadline.")

    output_wb = Workbook()
    detail_ws = output_wb.active
    detail_ws.title = "Overdue Sanctions"

    output_headers = REQUIRED_OUTPUT_FIELDS
    detail_ws.append(output_headers)

    records = []
    for row in source_ws.iter_rows(min_row=2, values_only=True):
        deadline_idx = columns["next_deadline"] - 1
        row_deadline = _parse_date_value(row[deadline_idx] if deadline_idx < len(row) else None)
        if row_deadline is None:
            continue
        if row_deadline < date_from or row_deadline > date_to:
            continue

        mapped = _build_output_row(row, columns)
        if not any(value for value in mapped.values()):
            continue
        records.append(mapped)
        detail_ws.append([mapped[header] for header in output_headers])

    if not records:
        raise OverdueSanctionsError("No matching rows were found for the selected date range.")

    for cell in detail_ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")

    sanctions_col = output_headers.index("Assigned Sanctions") + 1
    email_col = output_headers.index("Email") + 1
    deadline_col = output_headers.index("Next Deadline") + 1
    hold_col = output_headers.index("Hold") + 1
    send_email_col = output_headers.index("Send an Email") + 1
    hold_fill = PatternFill(fill_type="solid", fgColor="F8D7DA")
    send_email_row_fill = PatternFill(fill_type="solid", fgColor="FFF3CD")

    for row_idx in range(2, detail_ws.max_row + 1):
        sanctions_cell = detail_ws.cell(row=row_idx, column=sanctions_col)
        sanctions_cell.alignment = Alignment(vertical="top", wrap_text=True)

        email_cell = detail_ws.cell(row=row_idx, column=email_col)
        if _is_missing(email_cell.value):
            email_cell.fill = MISSING_EMAIL_FILL

        deadline_cell = detail_ws.cell(row=row_idx, column=deadline_col)
        if _is_missing(deadline_cell.value):
            deadline_cell.fill = MISSING_DEADLINE_FILL

        hold_cell = detail_ws.cell(row=row_idx, column=hold_col)
        send_email_cell = detail_ws.cell(row=row_idx, column=send_email_col)

        if str(send_email_cell.value).strip().lower() == "yes":
            for col_idx in range(1, detail_ws.max_column + 1):
                detail_ws.cell(row=row_idx, column=col_idx).fill = send_email_row_fill

        if str(hold_cell.value).strip().lower() == "yes":
            hold_cell.fill = hold_fill

    detail_ws.freeze_panes = "A2"
    _set_column_widths(detail_ws)
    _append_summary_sheet(output_wb, records)

    stream = BytesIO()
    output_wb.save(stream)
    preview_columns = output_headers
    preview_rows = records[:PREVIEW_ROW_LIMIT]
    return ProcessedOverdueSanctionsResult(
        total_rows=len(records),
        preview_columns=preview_columns,
        preview_rows=preview_rows,
        workbook_content=stream.getvalue(),
    )
